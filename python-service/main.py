"""
Spec Analyzer Python Service
Deployed on Render at submittal4subs.onrender.com

Endpoints:
- POST /upload - Upload PDF, store in R2, create spec record
- POST /parse/{spec_id} - Parse PDF into pages with section tags
- POST /analyze/{spec_id} - Run AI analysis on a division
- GET /spec/{spec_id}/divisions - Get divisions found in a spec
- POST /reclassify - Reclassify emails using Senior PM AI
- GET /email-stats - Get email classification statistics
"""

import asyncio
import os
import sys
from typing import List, Optional
from uuid import uuid4

from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

load_dotenv()
load_dotenv(dotenv_path="../.env")


def get_env(key: str) -> str:
    return os.getenv(key) or os.getenv(f"VITE_{key}") or ""


if not os.getenv("SUPABASE_URL"):
    os.environ["SUPABASE_URL"] = get_env("SUPABASE_URL")
if not os.getenv("SUPABASE_SERVICE_KEY"):
    os.environ["SUPABASE_SERVICE_KEY"] = get_env("SUPABASE_SERVICE_KEY") or get_env(
        "SUPABASE_ANON_KEY"
    )
if not os.getenv("GEMINI_API_KEY"):
    os.environ["GEMINI_API_KEY"] = get_env("GEMINI_API_KEY")
if not os.getenv("OPENAI_API_KEY"):
    os.environ["OPENAI_API_KEY"] = get_env("OPENAI_API_KEY")

from analyzer import (
    TRADE_CONFIGS,
    analyze_division_by_section,
    run_full_analysis,
    should_use_section_analysis,
)
from db import (
    create_spec,
    delete_divisions,
    delete_job,
    delete_pages,
    delete_tiles,
    get_all_analyses,
    get_analysis,
    get_division_summary,
    get_pages_by_division,
    get_pages_by_section,
    get_related_sections,
    get_sections_for_division,
    get_spec,
    insert_analysis,
    insert_pages_batch,
    update_spec_status,
)
from email_classifier import reclassify_emails
from parser import parse_spec
from storage import (
    delete_submittal_file,
    download_pdf,
    download_submittal_file,
    upload_pdf,
    upload_submittal_file,
)

app = FastAPI(
    title="Spec Analyzer API",
    description="PDF spec parsing, AI analysis, and email classification service",
    version="3.1.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

print("[BOOT] Spec Analyzer Service v3.1 (Page-Level + Email Classifier)")
print(f"[BOOT] SUPABASE_URL: {'OK' if os.getenv('SUPABASE_URL') else 'MISSING'}")
print(
    f"[BOOT] SUPABASE_SERVICE_KEY: {'OK' if os.getenv('SUPABASE_SERVICE_KEY') else 'MISSING'}"
)
print(f"[BOOT] R2_ACCOUNT_ID: {'OK' if os.getenv('R2_ACCOUNT_ID') else 'MISSING'}")
print(f"[BOOT] GEMINI_API_KEY: {'OK' if os.getenv('GEMINI_API_KEY') else 'MISSING'}")
print(f"[BOOT] OPENAI_API_KEY: {'OK' if os.getenv('OPENAI_API_KEY') else 'MISSING'}")
print(
    f"[BOOT] ANTHROPIC_API_KEY: {'OK' if os.getenv('ANTHROPIC_API_KEY') else 'MISSING'}"
)
print(f"[BOOT] MAIL4SUBS_URL: {'OK' if os.getenv('MAIL4SUBS_URL') else 'MISSING'}")


# ═══════════════════════════════════════════════════════════════
# REQUEST/RESPONSE MODELS
# ═══════════════════════════════════════════════════════════════


class UploadResponse(BaseModel):
    spec_id: str
    r2_key: str
    original_name: str
    status: str


class ParseResponse(BaseModel):
    spec_id: str
    status: str
    page_count: int
    division_count: int
    divisions: list
    toc_found: bool = False
    classification_stats: dict = None


class AnalyzeRequest(BaseModel):
    division: str
    include_contract_terms: bool = True
    project_name: Optional[str] = None
    related_sections: Optional[List[str]] = None


class AnalyzeResponse(BaseModel):
    spec_id: str
    division: str
    pages_analyzed: int
    cross_refs_included: int
    analysis: dict
    processing_time_ms: int


class ReclassifyRequest(BaseModel):
    limit: int = 50
    only_unclassified: bool = False
    category_filter: Optional[str] = None


class ReclassifyResponse(BaseModel):
    status: str
    processed: int
    results: dict
    changes: list = []
    scope_alerts: list = []
    money_alerts: list = []


# ═══════════════════════════════════════════════════════════════
# HEALTH CHECK
# ═══════════════════════════════════════════════════════════════


@app.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "version": "3.1.0",
        "services": {
            "supabase": bool(os.getenv("SUPABASE_URL")),
            "r2": bool(os.getenv("R2_ACCOUNT_ID")),
            "gemini": bool(os.getenv("GEMINI_API_KEY")),
            "openai": bool(os.getenv("OPENAI_API_KEY")),
            "anthropic": bool(os.getenv("ANTHROPIC_API_KEY")),
            "mail4subs": bool(os.getenv("MAIL4SUBS_URL")),
        },
    }


# ═══════════════════════════════════════════════════════════════
# EMAIL RECLASSIFICATION ENDPOINTS
# ═══════════════════════════════════════════════════════════════


@app.post("/reclassify", response_model=ReclassifyResponse)
async def reclassify_emails_endpoint(request: ReclassifyRequest):
    """Reclassify emails using Claude Sonnet as Senior PM."""
    print(f"[RECLASSIFY] Request: limit={request.limit}")

    try:
        result = await reclassify_emails(
            limit=request.limit,
            only_unclassified=request.only_unclassified,
            category_filter=request.category_filter,
        )
        return ReclassifyResponse(
            status=result.get("status", "complete"),
            processed=result.get("processed", 0),
            results=result.get("results", {}),
            changes=result.get("changes", []),
            scope_alerts=result.get("scope_alerts", []),
            money_alerts=result.get("money_alerts", []),
        )
    except Exception as e:
        print(f"[RECLASSIFY] ERROR: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/email-stats")
async def get_email_stats():
    """Get email classification statistics from Mail4Subs."""
    import httpx

    MAIL4SUBS_URL = os.getenv(
        "MAIL4SUBS_URL", "https://qiyonvhubpevqrrisqdl.supabase.co"
    )
    MAIL4SUBS_KEY = os.getenv("MAIL4SUBS_SERVICE_KEY") or os.getenv("MAIL4SUBS_KEY", "")

    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.get(
            f"{MAIL4SUBS_URL}/rest/v1/emails?select=category",
            headers={
                "apikey": MAIL4SUBS_KEY,
                "Authorization": f"Bearer {MAIL4SUBS_KEY}",
            },
        )

        if response.status_code != 200:
            raise HTTPException(status_code=500, detail="Failed to fetch stats")

        emails = response.json()
        stats = {
            "bid_invite": 0,
            "bid_result": 0,
            "quote_request": 0,
            "rfi": 0,
            "change_order": 0,
            "schedule": 0,
            "submittal": 0,
            "urgent": 0,
            "action": 0,
            "fyi": 0,
            "noise": 0,
            "bid": 0,
            "quote": 0,
            "unclassified": 0,
            "total": len(emails),
        }

        for email in emails:
            cat = email.get("category")
            if cat and cat in stats:
                stats[cat] += 1
            elif not cat:
                stats["unclassified"] += 1

        return stats


# ═══════════════════════════════════════════════════════════════
# SPEC UPLOAD/PARSE/ANALYZE ENDPOINTS
# ═══════════════════════════════════════════════════════════════


@app.post("/upload", response_model=UploadResponse)
async def upload_spec(
    file: UploadFile = File(...), user_id: str = Form(...), job_id: str = Form(...)
):
    """Upload a PDF specification."""
    print(f"[UPLOAD] User: {user_id}, Job: {job_id}, File: {file.filename}")

    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="File must be a PDF")

    try:
        pdf_bytes = await file.read()
        spec_id = str(uuid4())
        r2_key = upload_pdf(user_id, job_id, spec_id, pdf_bytes)
        spec = create_spec(
            user_id=user_id, job_id=job_id, r2_key=r2_key, original_name=file.filename
        )

        return UploadResponse(
            spec_id=spec["id"],
            r2_key=r2_key,
            original_name=file.filename,
            status="uploaded",
        )
    except Exception as e:
        print(f"[UPLOAD] ERROR: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/parse/{spec_id}", response_model=ParseResponse)
async def parse_spec_endpoint(spec_id: str):
    """Parse a PDF specification into pages with section tags."""
    print(f"[PARSE] Parsing spec: {spec_id}")

    spec = get_spec(spec_id)
    if not spec:
        raise HTTPException(status_code=404, detail="Spec not found")

    try:
        update_spec_status(spec_id, "processing")
        pdf_bytes = download_pdf(spec["r2_key"])

        delete_pages(spec_id)
        delete_divisions(spec_id)
        delete_tiles(spec_id)

        result = parse_spec(pdf_bytes, spec_id)

        if result["pages"]:
            insert_pages_batch(result["pages"])

        update_spec_status(spec_id, "ready", result["page_count"])

        division_list = []
        for div_code in sorted(result["division_summary"].keys()):
            info = result["division_summary"][div_code]
            pages = info["pages"]
            division_list.append(
                {
                    "code": div_code,
                    "page_count": info["count"],
                    "sections": info["sections"],
                    "section_count": len(info["sections"]),
                    "page_range": f"{min(pages)}-{max(pages)}" if pages else None,
                }
            )

        return ParseResponse(
            spec_id=spec_id,
            status="ready",
            page_count=result["page_count"],
            division_count=len(result["divisions"]),
            divisions=division_list,
            toc_found=result.get("toc_found", False),
            classification_stats=result.get("classification_stats"),
        )
    except Exception as e:
        print(f"[PARSE] ERROR: {e}")
        import traceback

        traceback.print_exc()
        update_spec_status(spec_id, "failed")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/analyze/{spec_id}", response_model=AnalyzeResponse)
async def analyze_spec_endpoint(spec_id: str, request: AnalyzeRequest):
    """Run AI analysis on a specific division."""
    division = request.division.zfill(2)
    print(f"[ANALYZE] Spec: {spec_id}, Division: {division}")

    spec = get_spec(spec_id)
    if not spec:
        raise HTTPException(status_code=404, detail="Spec not found")
    if spec["status"] != "ready":
        raise HTTPException(status_code=400, detail=f"Spec not ready: {spec['status']}")

    trade = None
    for trade_name, config in TRADE_CONFIGS.items():
        if config["division"] == division:
            trade = trade_name
            break
    if not trade:
        trade = "general"

    try:
        division_pages = get_pages_by_division(spec_id, division)
        if not division_pages:
            raise HTTPException(
                status_code=404, detail=f"No pages for Division {division}"
            )

        sections = get_sections_for_division(spec_id, division)
        section_count = len(sections)
        page_count = len(division_pages)

        use_section_analysis = should_use_section_analysis(page_count, section_count)

        if use_section_analysis:
            contract_analysis = None
            contract_summary_text = None
            if request.include_contract_terms:
                div00_pages = get_pages_by_division(spec_id, "00")
                div01_pages = get_pages_by_division(spec_id, "01")
                contract_pages = div00_pages + div01_pages
                if contract_pages:
                    from analyzer import analyze_contract_terms

                    div01_text = "\n\n".join(
                        [
                            f"--- Page {p['page_number']} ---\n{p['content']}"
                            for p in sorted(
                                contract_pages, key=lambda x: x["page_number"]
                            )
                        ]
                    )
                    contract_analysis = await analyze_contract_terms(
                        div01_text, request.project_name
                    )
                    contract_summary_text = contract_analysis.get("summary", "")

            related_section_count = 0
            if request.related_sections:
                for section_num in request.related_sections:
                    section_pages = get_pages_by_section(spec_id, section_num)
                    if section_pages:
                        related_content = "\n\n".join(
                            [
                                f"--- Page {p['page_number']} ---\n{p['content']}"
                                for p in sorted(
                                    section_pages, key=lambda x: x["page_number"]
                                )
                            ]
                        )
                        sections.append(
                            {
                                "section_number": section_num,
                                "title": f"Related Section {section_num}",
                                "page_count": len(section_pages),
                                "content": related_content,
                            }
                        )
                        related_section_count += 1

            analysis_result = await analyze_division_by_section(
                sections=sections,
                trade=trade,
                division=division,
                project_name=request.project_name,
                contract_summary=contract_summary_text,
            )
            if contract_analysis:
                analysis_result["contract_analysis"] = contract_analysis
            cross_ref_count = related_section_count
        else:
            division_text = "\n\n".join(
                [
                    f"--- Page {p['page_number']} (Section {p['section_number'] or 'unknown'}) ---\n{p['content']}"
                    for p in sorted(division_pages, key=lambda x: x["page_number"])
                ]
            )

            related_section_pages = []
            if request.related_sections:
                for section_num in request.related_sections:
                    section_pages = get_pages_by_section(spec_id, section_num)
                    related_section_pages.extend(section_pages)
                if related_section_pages:
                    related_text = "\n\n".join(
                        [
                            f"--- Page {p['page_number']} ---\n{p['content']}"
                            for p in sorted(
                                related_section_pages, key=lambda x: x["page_number"]
                            )
                        ]
                    )
                    division_text += f"\n\n{'=' * 60}\nRELATED SECTIONS\n{'=' * 60}\n\n{related_text}"

            cross_ref_count = len(related_section_pages)

            div01_text = None
            if request.include_contract_terms:
                div00_pages = get_pages_by_division(spec_id, "00")
                div01_pages = get_pages_by_division(spec_id, "01")
                contract_pages = div00_pages + div01_pages
                if contract_pages:
                    div01_text = "\n\n".join(
                        [
                            f"--- Page {p['page_number']} ---\n{p['content']}"
                            for p in sorted(
                                contract_pages, key=lambda x: x["page_number"]
                            )
                        ]
                    )

            analysis_result = await run_full_analysis(
                division_text=division_text,
                div01_text=div01_text,
                trade=trade,
                project_name=request.project_name,
            )

        insert_analysis(
            spec_id=spec_id,
            job_id=spec["job_id"],
            division_code=division,
            analysis_type="section_by_section" if use_section_analysis else "full",
            result=analysis_result,
            processing_time_ms=analysis_result["processing_time_ms"],
        )

        return AnalyzeResponse(
            spec_id=spec_id,
            division=division,
            pages_analyzed=len(division_pages),
            cross_refs_included=cross_ref_count,
            analysis=analysis_result,
            processing_time_ms=analysis_result["processing_time_ms"],
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"[ANALYZE] ERROR: {e}")
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ═══════════════════════════════════════════════════════════════
# SPEC DATA RETRIEVAL ENDPOINTS
# ═══════════════════════════════════════════════════════════════


@app.get("/spec/{spec_id}/analyses")
async def get_spec_analyses(spec_id: str):
    spec = get_spec(spec_id)
    if not spec:
        raise HTTPException(status_code=404, detail="Spec not found")
    analyses = get_all_analyses(spec_id)
    return {
        "spec_id": spec_id,
        "count": len(analyses),
        "analyses": [
            {
                "id": a["id"],
                "division_code": a["division_code"],
                "analysis_type": a["analysis_type"],
                "created_at": a["created_at"],
                "processing_time_ms": a["processing_time_ms"],
                "result": a["result"],
            }
            for a in analyses
        ],
    }


@app.get("/spec/{spec_id}/analysis/{division}")
async def get_division_analysis(spec_id: str, division: str):
    division = division.zfill(2)
    spec = get_spec(spec_id)
    if not spec:
        raise HTTPException(status_code=404, detail="Spec not found")
    analysis = get_analysis(spec_id, division)
    if not analysis:
        raise HTTPException(
            status_code=404, detail=f"No analysis for Division {division}"
        )
    return {
        "spec_id": spec_id,
        "division_code": division,
        "id": analysis["id"],
        "created_at": analysis["created_at"],
        "processing_time_ms": analysis["processing_time_ms"],
        "result": analysis["result"],
    }


@app.get("/spec/{spec_id}/divisions")
async def get_spec_divisions(spec_id: str):
    spec = get_spec(spec_id)
    if not spec:
        raise HTTPException(status_code=404, detail="Spec not found")
    divisions = get_division_summary(spec_id)
    return {
        "spec_id": spec_id,
        "status": spec["status"],
        "page_count": spec.get("page_count"),
        "divisions": [
            {
                "code": d["division_code"],
                "page_count": d["page_count"],
                "sections": d.get("sections", []),
                "page_range": d.get("page_range"),
            }
            for d in divisions
        ],
    }


@app.get("/spec/{spec_id}/division/{division}/related")
async def get_division_related_sections(spec_id: str, division: str):
    spec = get_spec(spec_id)
    if not spec:
        raise HTTPException(status_code=404, detail="Spec not found")
    related = get_related_sections(spec_id, division)
    return {"spec_id": spec_id, "division": division, "related_sections": related}


@app.delete("/job/{job_id}")
async def delete_job_endpoint(job_id: str, user_id: str):
    if not user_id:
        raise HTTPException(status_code=400, detail="user_id is required")
    try:
        success = delete_job(job_id, user_id)
        if not success:
            raise HTTPException(status_code=404, detail="Job not found")
        return {"status": "deleted", "job_id": job_id}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ═══════════════════════════════════════════════════════════════
# SUBMITTAL FILE ENDPOINTS
# ═══════════════════════════════════════════════════════════════


class SubmittalUploadResponse(BaseModel):
    r2_key: str
    file_name: str
    file_size: int


class SubmittalDeleteRequest(BaseModel):
    r2_key: str


@app.post("/submittal/upload", response_model=SubmittalUploadResponse)
async def upload_submittal(file: UploadFile = File(...), item_id: str = Form(...)):
    allowed_extensions = {
        ".pdf",
        ".doc",
        ".docx",
        ".xls",
        ".xlsx",
        ".rtf",
        ".txt",
        ".jpg",
        ".jpeg",
        ".png",
        ".gif",
        ".tif",
        ".tiff",
        ".bmp",
        ".csv",
        ".ppt",
        ".pptx",
        ".odt",
        ".ods",
        ".odp",
    }
    ext = os.path.splitext(file.filename.lower())[1]
    if ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail=f"File type '{ext}' not supported")

    try:
        file_bytes = await file.read()
        r2_key = upload_submittal_file(item_id, file.filename, file_bytes)
        return SubmittalUploadResponse(
            r2_key=r2_key, file_name=file.filename, file_size=len(file_bytes)
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/submittal/download/{r2_key:path}")
async def download_submittal(r2_key: str):
    mime_types = {
        ".pdf": "application/pdf",
        ".doc": "application/msword",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".xls": "application/vnd.ms-excel",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
    }
    try:
        file_bytes = download_submittal_file(r2_key)
        filename = r2_key.split("/")[-1]
        ext = os.path.splitext(filename.lower())[1]
        content_type = mime_types.get(ext, "application/octet-stream")
        from fastapi.responses import Response

        return Response(
            content=file_bytes,
            media_type=content_type,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=404, detail="File not found")


@app.get("/submittal/file/{r2_key:path}")
async def get_submittal_file(r2_key: str):
    try:
        pdf_bytes = download_submittal_file(r2_key)
        from fastapi.responses import Response

        return Response(content=pdf_bytes, media_type="application/pdf")
    except Exception as e:
        raise HTTPException(status_code=404, detail="File not found")


@app.post("/submittal/delete")
async def delete_submittal(request: SubmittalDeleteRequest):
    try:
        success = delete_submittal_file(request.r2_key)
        if success:
            return {"status": "deleted", "r2_key": request.r2_key}
        raise HTTPException(status_code=500, detail="Failed to delete file")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ═══════════════════════════════════════════════════════════════
# SUBMITTAL AI EXTRACTION
# ═══════════════════════════════════════════════════════════════


class ExtractSubmittalsRequest(BaseModel):
    text: str


class SubmittalItem(BaseModel):
    spec_section: str
    description: str
    manufacturer: str


class ExtractSubmittalsResponse(BaseModel):
    items: List[SubmittalItem]
    error: Optional[str] = None


@app.post("/extract-submittals", response_model=ExtractSubmittalsResponse)
async def extract_submittals(request: ExtractSubmittalsRequest):
    import json

    import httpx

    if not request.text:
        return ExtractSubmittalsResponse(items=[], error="No text provided")

    gemini_api_key = os.getenv("GEMINI_API_KEY")
    if not gemini_api_key:
        return ExtractSubmittalsResponse(items=[], error="AI service not configured")

    prompt = f"""Extract ONLY physical materials and products requiring submittals from this construction spec analysis.
Return ONLY a valid JSON array, no other text:
[{{"spec_section": "04 20 00", "description": "Item name", "manufacturer": "Manufacturer or empty string"}}]

INCLUDE: Physical materials, products with manufacturers, items from "Quote These Items"
DO NOT INCLUDE: Schedules, certificates, forms, closeout docs, test reports, meeting minutes

Analysis text:
{request.text}"""

    try:
        gemini_url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent"
        async with httpx.AsyncClient(timeout=60.0) as client:
            response = await client.post(
                f"{gemini_url}?key={gemini_api_key}",
                json={
                    "contents": [{"parts": [{"text": prompt}]}],
                    "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048},
                },
            )
            if response.status_code != 200:
                return ExtractSubmittalsResponse(
                    items=[], error=f"AI API error: {response.status_code}"
                )

            result = response.json()
            result_text = (
                result.get("candidates", [{}])[0]
                .get("content", {})
                .get("parts", [{}])[0]
                .get("text", "")
                .strip()
            )

            if result_text.startswith("```"):
                result_text = result_text.split("\n", 1)[1].rsplit("```", 1)[0]
            if result_text.startswith("json"):
                result_text = result_text[4:].strip()

            items = json.loads(result_text)
            valid_items = [
                SubmittalItem(
                    spec_section=str(item.get("spec_section", "")),
                    description=str(item.get("description", "Unknown")),
                    manufacturer=str(item.get("manufacturer", "")),
                )
                for item in items
            ]
            return ExtractSubmittalsResponse(items=valid_items)
    except json.JSONDecodeError:
        return ExtractSubmittalsResponse(items=[], error="Failed to parse AI response")
    except Exception as e:
        return ExtractSubmittalsResponse(items=[], error=str(e))


# ═══════════════════════════════════════════════════════════════
# FILE CONVERSION
# ═══════════════════════════════════════════════════════════════


@app.get("/submittal/file-as-pdf/{r2_key:path}")
async def get_submittal_file_as_pdf(r2_key: str):
    filename = r2_key.split("/")[-1]
    ext = os.path.splitext(filename.lower())[1]
    convertible_extensions = {
        ".doc",
        ".docx",
        ".rtf",
        ".txt",
        ".odt",
        ".xls",
        ".xlsx",
        ".ods",
        ".csv",
        ".ppt",
        ".pptx",
        ".odp",
    }
    image_extensions = {".jpg", ".jpeg", ".png", ".gif", ".tif", ".tiff", ".bmp"}

    try:
        file_bytes = download_submittal_file(r2_key)

        if ext == ".pdf":
            from fastapi.responses import Response

            return Response(content=file_bytes, media_type="application/pdf")

        if ext in image_extensions:
            pdf_bytes = convert_image_to_pdf(file_bytes, ext)
            if pdf_bytes:
                from fastapi.responses import Response

                return Response(content=pdf_bytes, media_type="application/pdf")
            raise HTTPException(status_code=500, detail="Image conversion failed")

        if ext in convertible_extensions:
            pdf_bytes = convert_document_to_pdf(file_bytes, filename)
            if pdf_bytes:
                from fastapi.responses import Response

                return Response(content=pdf_bytes, media_type="application/pdf")
            raise HTTPException(status_code=500, detail="Document conversion failed")

        raise HTTPException(
            status_code=400, detail=f"File type {ext} cannot be converted to PDF"
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def convert_image_to_pdf(image_bytes: bytes, ext: str) -> Optional[bytes]:
    try:
        from io import BytesIO

        from PIL import Image

        img = Image.open(BytesIO(image_bytes))
        if img.mode in ("RGBA", "LA", "P"):
            background = Image.new("RGB", img.size, (255, 255, 255))
            if img.mode == "P":
                img = img.convert("RGBA")
            background.paste(img, mask=img.split()[-1] if img.mode == "RGBA" else None)
            img = background
        elif img.mode != "RGB":
            img = img.convert("RGB")
        pdf_buffer = BytesIO()
        img.save(pdf_buffer, format="PDF", resolution=100.0)
        pdf_buffer.seek(0)
        return pdf_buffer.getvalue()
    except Exception as e:
        print(f"[CONVERT] Image error: {e}")
        return None


def convert_document_to_pdf(doc_bytes: bytes, filename: str) -> Optional[bytes]:
    ext = os.path.splitext(filename.lower())[1]
    if ext in (".docx", ".doc"):
        return convert_docx_to_pdf(doc_bytes)
    elif ext in (".xlsx", ".xls", ".csv"):
        return convert_excel_to_pdf(doc_bytes, ext)
    elif ext in (".txt", ".rtf"):
        return convert_text_to_pdf(doc_bytes, ext)
    return None


def convert_docx_to_pdf(doc_bytes: bytes) -> Optional[bytes]:
    try:
        from io import BytesIO

        from docx import Document
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

        doc = Document(BytesIO(doc_bytes))
        pdf_buffer = BytesIO()
        pdf_doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=letter,
            leftMargin=inch,
            rightMargin=inch,
            topMargin=inch,
            bottomMargin=inch,
        )
        styles = getSampleStyleSheet()
        story = []
        for para in doc.paragraphs:
            if para.text.strip():
                text = (
                    para.text.replace("&", "&amp;")
                    .replace("<", "&lt;")
                    .replace(">", "&gt;")
                )
                story.append(
                    Paragraph(
                        text,
                        styles["Heading1"]
                        if para.style.name.startswith("Heading")
                        else styles["Normal"],
                    )
                )
                story.append(Spacer(1, 6))
        if not story:
            story.append(Paragraph("(Empty document)", styles["Normal"]))
        pdf_doc.build(story)
        pdf_buffer.seek(0)
        return pdf_buffer.getvalue()
    except Exception as e:
        print(f"[CONVERT] DOCX error: {e}")
        return None


def convert_excel_to_pdf(doc_bytes: bytes, ext: str) -> Optional[bytes]:
    try:
        from io import BytesIO, StringIO

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

        if ext == ".csv":
            import csv

            content = doc_bytes.decode("utf-8", errors="replace")
            reader = csv.reader(StringIO(content))
            data = list(reader)
        else:
            from openpyxl import load_workbook

            wb = load_workbook(BytesIO(doc_bytes), data_only=True)
            ws = wb.active
            data = []
            for row in ws.iter_rows(max_row=100, max_col=20):
                row_data = [
                    str(cell.value) if cell.value is not None else "" for cell in row
                ]
                if any(row_data):
                    data.append(row_data)
        if not data:
            data = [["(Empty spreadsheet)"]]
        pdf_buffer = BytesIO()
        pdf_doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(letter))
        table = Table(data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ]
            )
        )
        pdf_doc.build([table])
        pdf_buffer.seek(0)
        return pdf_buffer.getvalue()
    except Exception as e:
        print(f"[CONVERT] Excel error: {e}")
        return None


def convert_text_to_pdf(doc_bytes: bytes, ext: str) -> Optional[bytes]:
    try:
        from io import BytesIO

        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

        try:
            text = doc_bytes.decode("utf-8")
        except UnicodeDecodeError:
            text = doc_bytes.decode("latin-1", errors="replace")
        if ext == ".rtf":
            import re

            text = re.sub(r"\\[a-z]+\d*\s?", "", text)
            text = re.sub(r"[{}]", "", text)
        pdf_buffer = BytesIO()
        pdf_doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=letter,
            leftMargin=inch,
            rightMargin=inch,
            topMargin=inch,
            bottomMargin=inch,
        )
        styles = getSampleStyleSheet()
        story = []
        for para in text.split("\n\n"):
            para = para.strip()
            if para:
                para = (
                    para.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                )
                story.append(Paragraph(para, styles["Normal"]))
                story.append(Spacer(1, 6))
        if not story:
            story.append(Paragraph("(Empty document)", styles["Normal"]))
        pdf_doc.build(story)
        pdf_buffer.seek(0)
        return pdf_buffer.getvalue()
    except Exception as e:
        print(f"[CONVERT] Text error: {e}")
        return None


# ═══════════════════════════════════════════════════════════════
# RUN SERVER
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import uvicorn

    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
